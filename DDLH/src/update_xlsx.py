#-*-coding:UTF-8-*-
'''
Created on 2018��2��10��-����3:04:23
author: Gary-W
'''

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

import os

class OrderItem:
    def __init__(self):
        self.item_id = None         # �Ϻ�
        self.order_id = None        # ������
        self.income_num = 0        # ��������
    
    def __str__(self):
        return "ID:%s ORD:%s N:%d" % (self.item_id, self.order_id, self.income_num)
    
    
class OrderDataSheet:
    """ ��ȡ�������XLSX��񣬻�ȡ��������
    """
    def __init__(self,xlsx_path, sheetname="325"):  #zhijia
        """ �������е�xls�����ƶ����е�sheet
        """
        self.xlsx_path = xlsx_path
        self.workbook = load_workbook(xlsx_path)
        self.main_sheet = self.workbook.get_sheet_by_name(sheetname)

        # ��ȡ������
        self.header_rows_offset = 1    # ������ռ����
        self.sample_num = self._get_sample_num()
        self.item_list = []
        self.get_items()
        
    def _get_sample_num(self):
        """ ͳ���������Ч����
        """
        rows = len(self.main_sheet.rows)
        for i in range(rows,0,-1):
            if self.main_sheet["B"+str(i)].value != None:
                sample_num = i - 1
                break
        return sample_num

    def get_items(self):
        """ ��ñ������е�����
        """
        for row_id in range(self.sample_num):
            orditem = OrderItem()
            orditem.item_id = str(self.main_sheet["E"+str(row_id+2)].value)
            orditem.order_id = str(self.main_sheet["I"+str(row_id+2)].value)
            orditem.income_num = int(self.main_sheet["G"+str(row_id+2)].value)
            self.item_list.append(orditem)


class TrackItem:
    def __init__(self):
        self.item_id = None         # �Ϻ�
        self.order_id = None        # ������
        self.target_num = 0         # Ŀ��������
        self.income_num = 0         # ���������
        self.rest_num = 0           # δ������
        self.mat_row = 0            # �û������ڵ���

    def __str__(self):
        return "%s %s %d %d" % (self.item_id, self.order_id, self.target_num, self.rest_num)
    
class TrackDataSheet:
    """ д�뵽�����XLSX��񣬸�����������
    """
    def __init__(self,xlsx_path, sheetname="Sheet1"):
        """ �������е�xls�����ƶ����е�sheet
        """
        self.xlsx_path = xlsx_path
        self.workbook = load_workbook(xlsx_path)
        self.main_sheet = self.workbook.get_sheet_by_name(sheetname)

        # ��ȡ������
        self.header_rows_offset = 2    # ������ռ����
        self.sample_num = self._get_sample_num()
        self.item_id_col = 1
        self.ord_id_col = 2

        self.item_table = {}
        self._get_item_table()

    def _get_sample_num(self):
        """ ͳ���������Ч����
        """
        rows = len(self.main_sheet.rows)
        for i in range(rows,0,-1):
            if self.main_sheet["B"+str(i)].value != None:
                sample_num = i - 2
                break
        return sample_num

    def _get_item_table(self):
        latest_item_id = None
        for row_id in range(self.sample_num):
            item_id = self.main_sheet["A"+str(row_id+3)].value
            order_id = self.main_sheet["B"+str(row_id+3)].value
            
            if item_id is not None:
                self.item_table[item_id] = {}
                latest_item_id = item_id

            latest_item_id = str(latest_item_id)
            order_id = str(order_id)
            
            # ��д �Ϻţ������ţ�������δ�������� ��Ӧ�ı���id
            track_item = TrackItem()
            track_item.item_id = latest_item_id
            track_item.order_id = order_id
            track_item.mat_row = row_id+3
            
            track_item.target_num = int(self.main_sheet["F"+str(row_id+3)].value)
            track_item.rest_num = int(self.main_sheet["G"+str(row_id+3)].value)
            track_item.income_num = track_item.target_num - track_item.rest_num
            self.item_table[latest_item_id][order_id] = track_item
    
    
    def load(self, insert_col, income_items):
        order_col = insert_col
        income_col = insert_col + 1
        rest_col = 7    # δ����������
        err_msg_list = []
        align = Alignment(horizontal='right', vertical='center')
        for i, income_item in enumerate(income_items):
            if income_item.item_id in self.item_table:
                item_id = income_item.item_id
                order_id = income_item.order_id
                income_num = income_item.income_num
                if order_id not in self.item_table[item_id]:
                    err_msg = "ItemID: %s ORD: %s not found: %s" % (item_id, order_id, income_item)
                    err_msg_list.append(err_msg)
                else:
                    # ����ƥ��������Ϣ
                    insert_row = self.item_table[item_id][order_id].mat_row
                    
                    # �������ź��������д���Ӧ�ĵ�Ԫ��
                    order_cell = self.main_sheet.cell(row = insert_row, column = order_col)
                    income_cell = self.main_sheet.cell(row = insert_row, column = income_col)
                    rest_cell = self.main_sheet.cell(row = insert_row, column = rest_col)
                    
                    # ����δ������
                    rest_cell.value = rest_cell.value - income_num
                    # ���µ�ǰ�ѽ�����
                    if income_cell.value == None:
                        income_cell.value = income_num
                    else:
                        income_cell.value += income_num
                    # ���¶�����
                    order_cell.value = int(order_id)
                    
                    order_cell.alignment = align
                    income_cell.alignment = align
            else:
                err_msg = "ItemID: %s not found: %s" % (income_item.item_id, income_item)
                err_msg_list.append(err_msg)
        print "update:",self.xlsx_path
        self.workbook.save(self.xlsx_path)

        err_log_path = self.xlsx_path[:-5] + "_err.txt"
        with open(err_log_path, "w") as f:
            for err_msg in err_msg_list:
                f.write(err_msg+"\n")



def main():
    
#     input_xlsx_path = r"D:\������ϸ.xlsx"
#     update_xlsx_path = r"D:\���϶���ִ�и��ٱ�.xlsx"
    input_xlsx_path = raw_input("����ϸ ·��: ")
    sheetname = raw_input("������ϸ����: ")
    update_xlsx_path = raw_input("���϶���ִ�и��ٱ� ·��: ")
    insert_col = raw_input("������ٱ���µ���: ")

    if not os.path.exists(input_xlsx_path):
        print "�Ҳ���",input_xlsx_path
        raw_input("���س��˳�")
        return
    
    if not os.path.exists(update_xlsx_path):
        print "�Ҳ���",update_xlsx_path
        raw_input("���س��˳�")
        return

    insert_col = int(insert_col)
    if insert_col < 10:
        print "�޷���Ӧ��", insert_col
        raw_input("���س��˳�")
        return

    ds_input = OrderDataSheet(input_xlsx_path, sheetname)

    ds_update = TrackDataSheet(update_xlsx_path)
    ds_update.load(insert_col, ds_input.item_list)
    raw_input("���³ɹ� ���س��˳�")
 



if __name__=="__main__":
    pass
    main()
